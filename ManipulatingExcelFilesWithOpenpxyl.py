# Manipulating Excel Files with openpxyl:

# Overview of openpxyl:

# The openpyxl library:
#       Free, open source Python library for working with Excel files (.xlsx)
#       Read, write, and manipulate the content of workbooks and worksheets
#       Userful for automating many common tasks:
#           Create spreadsheets from other data sources
#           Combine multiple data sources and generate formatted reports
#           Create charts and templates programmatically
#       Build and deploy powerful data processing pipelines and workflows
# Openpyxl is also used by the Pandas library to perform operations on Excel files


# Loading and exploring a workbook:

# Open, load, and explore workbook content 

'''
import openpyxl

filename = "FinancialSample.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(filename)

# Print basic information
print(f"Number of worksheets: {len(workbook.sheetnames)}")
for worksheet_name in workbook.sheetnames:
    worksheet = workbook[worksheet_name]
    print(f"\nWorksheet: {worksheet_name}")

# Explore each worksheet

    
# Get dimensions
    dimensions = worksheet.dimensions
    print(f"  - Dimensions: {dimensions}")

    print(f"Min row: {worksheet.min_row}")
    print(f"Max row: {worksheet.max_row}")
    print(f"Min column: {worksheet.min_column}")
    print(f"Max column: {worksheet.max_column}")

# Check if the worksheet is empty
    if worksheet.max_row == 1 and worksheet.max_column == 1:
        print(f"  - Worksheet is empty")
    else:
        cell = worksheet["A1"]
        print(f"  - Top=left cell calue: {cell.value}")
        cell = worksheet.cell(row=worksheet.max_row, column=worksheet.max_column)
        print(f"  - Bottom-right cell value: {cell.value}")
'''


# Creating a workbook:

# Create a new workbook with worksheets and add content 

'''
from openpyxl import Workbook
import datetime
import random

# Create a new workbook
wb = Workbook()

# Get the active worksheet and name it "First"
sheet = wb.active
sheet.title = "First"

# Add some data to the new sheet
sheet["A1"] = "Test Data"
sheet["B1"] = 123.4567
sheet["C1"] = datetime.datetime(2030, 4, 1)

# Use the cell() function to fill a row with values
for i in range(1, 11):
    sheet.cell(row=5, column=i).value = random.randint(1,50)

# Create a second worksheet
sheet2 = wb.create_sheet("Second")
sheet2.cell(row=2, column=2).value = "More Data"

# Use the append() function to add rows to the end of the sheet
sheet2.append(["One", "Two", "Three"])
sheet2.append(["One", "Two", "Three"])
sheet2.append(["One", "Two", "Three"])

# Save the workbook - values don't update until we do this!
wb.save("NewWorkbook.xlsx")
print("Workbook has been created successfully!")
'''


# Working with Content:

# Can get an entire column or row of cells by using either a letter or numberic index recpecivly

'''
import openpyxl
from openpyxl.comments import Comment
from collections import defaultdict


# Create a new workbook
filename = "FinancialSample.xlsx"

# Load the workbook
wb = openpyxl.load_workbook(filename)

# Get the active worksheet
sheet = wb.active

# Get entire column or row of cells
col = sheet["C"]
row = sheet[10]
print(f"{len(col)} cells in a column")
print(f"{len(row)} cells in a row")

# Get a range of cells
range = sheet["A2:B7"]
print(f"{len(range)} cells in range")
print(range)

# iterate over rows and columns
for col in sheet.iter_cols(min_row=2, max_row=3, min_col=2, max_col=5):
    for cell in col:
        print(cell.value)
counter = defaultdict(int)
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        counter[cell.value] +=1
print(counter)

# create a cell with a comment in it
cell = sheet["A1"]
cell.comment = Comment("This is a comment", "James Allen")

# save the workbook
wb.save("Content.xlsx")
'''


# Styling Cells:

# Manipulate cell content and styling 

'''
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import openpyxl.styles.numbers as opnumstyle
import datetime


# Create a new workbook
wb = Workbook()

# Get the active worksheet and name it "TestSheet"
sheet = wb.active
sheet.title = "First"

# Add some data to the new sheet
sheet["A1"] = "Test Data"
sheet["B1"] = 123.4567
sheet["C1"] = datetime.datetime(2030, 4, 1)

# Inspect the default styles of each cell
print(sheet["A1"].style)
print(sheet["B1"].number_format)
print(sheet["C1"].number_format)

# Use some built-in styles
# Moving on from these to custom syles down below
#sheet["A1"].style = "Title"
#sheet["B1"].style = "Calculation"
#sheet["C1"].style = "Accent2"
#sheet["B1"].number_format = opnumstyle.FORMAT_CURRENCY_USD_SIMPLE
#sheet["C1"].number_format = opnumstyle.FORMAT_DATE_DDMMYY

sheet.column_dimensions['A'].width = 30
sheet.row_dimensions[1].height = 50

# Create styles using Fonts and Colors
italic_font = Font(italic=True, size=16)
colored_text = Font(name="Courier New", size=20, color="000000FF")
centered_text = Alignment(horizontal="center", vertical="top")
border_side = Side(border_style="mediumDashed")
cell_border = Border(top=border_side, right=border_side, left=border_side, bottom=border_side)

sheet["A1"].font = italic_font
sheet["B1"].font = colored_text
sheet["B1"].alignment = centered_text
sheet["c1"].border = cell_border

# Save the workbook
wb.save("StyledCells.xlsx")
'''


# Applying Conditional Formatting:

# One of the great features of Excel is the ability to perform conditional formatting
# Which helps highlight specific data conditions when you have a large amount of data in a worksheet
# Openpyxl lets you add conditional formatting to Excel data with simple Python code

# Apply conditional formatting to a worksheet 

'''
import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle


filename = "FinancialSample.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(filename)
sheet = workbook["SalesData"]

# define the style to represent the formatting
red_color = "ffd2d2"
bold_text = Font(bold=True, color="00ff0000")
red_fill = PatternFill(bgColor=red_color, fill_type="solid")

diff_style = DifferentialStyle(font=bold_text, fill=red_fill)

# create a rule for the condition
rule = Rule(type="expression", dxf=diff_style, formula=["$L1<10000"])

# add the rule to the entire sheet
# Change from this: 
#dimensions = sheet.dimensions
#sheet.conditional_formatting.add(dimensions, rule)

# Changing to only highlight one row:
# This is new code:
dimensions = "L1:L701"
sheet.conditional_formatting.add(dimensions, rule)

workbook.save("CondFormat.xlsx")
print("Workbook created successfully!")
'''


# Adding Filters:

# When working with large amounts of data, it usually helps to be able to apply filters to the dataset 
#       so you can just focus on the parts that you care about at that moment. 
# In Excel, you can do this by applying filter controls to the columns in the dataset. Openpyxl lets you do this in Python

# Add column filters to a sheet

import csv
from openpyxl import Workbook


def read_csv_to_array(filename):
    # define the array that will hold the data
    data = []
    with open(filename, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            data.append(row)
    return data


# Read the data into an array of arrays
inventory_data = read_csv_to_array("Inventory.csv")

# Create a new workbook
wb = Workbook()

# Get the active worksheet and name it "TestSheet"
sheet = wb.active
sheet.title = "Inventory"

for row in inventory_data:
    sheet.append(row)

# Add the filters to the columns
filters = sheet.auto_filter
filters.ref = sheet.dimensions

wb.save("Inventory.xlsx")